"""v41 — Export des données de l'apprenant (droit à la portabilité, RGPD art. 20).

Un fichier Excel, structuré et lisible — pas un vidage de tables. Il contient
TOUT ce que Certifizer conserve sur la personne, et rien d'autre.

Le portrait (v40) reste le bel objet ; ceci est son pendant : exhaustif,
machine-lisible, réutilisable ailleurs. Les deux servent des maîtres différents,
et c'est voulu.
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INK = "0E1A2B"; AMBER = "E89A3C"; TEAL = "2E8C9E"; PAPER = "F5F8FA"; LINE = "DCE3EA"
FONT = "Arial"
_thin = Side(style="thin", color=LINE)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header(ws, row, cols, fill=INK):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = _border
    ws.row_dimensions[row].height = 24


def _rows(ws, start, data, widths=None):
    for r, line in enumerate(data, start):
        for c, v in enumerate(line, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT, size=9.5)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = _border
            if (r - start) % 2:
                cell.fill = PatternFill("solid", fgColor=PAPER)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def build_export(portrait: dict, name: str, email: str, cohort: str, lang: str = "fr") -> bytes:
    """Construit le classeur à partir du portrait déjà calculé (une seule source
    de vérité : le portrait et l'export ne peuvent pas se contredire)."""
    fr = lang != "en"
    wb = Workbook()

    # ---------- 1. Mon compte ----------
    ws = wb.active
    ws.title = "Mon compte" if fr else "My account"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Certifizer — " + ("Mes données" if fr else "My data")
    ws["A1"].font = Font(name=FONT, bold=True, size=15, color=INK)
    ws["A2"] = ("Export généré le " if fr else "Exported on ") + datetime.utcnow().strftime("%d/%m/%Y")
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="5E6E7F")

    _header(ws, 4, [("Information" if fr else "Information"), ("Valeur" if fr else "Value")], TEAL)
    info = [
        ("Nom" if fr else "Name", name or "—"),
        ("Email" if fr else "Email", email or ("(aucun — facultatif)" if fr else "(none — optional)")),
        ("Identifiant" if fr else "Identifier", portrait.get("learner_id", "")),
        ("Cohorte" if fr else "Cohort", cohort or "—"),
        ("Préparation" if fr else "Readiness", f"{round(portrait.get('readiness', 0) * 100)} %"),
        ("Domaines acquis" if fr else "Areas acquired", f"{portrait.get('acquired', 0)} / {portrait.get('total_areas', 0)}"),
        ("Réponses données" if fr else "Answers given", portrait.get("total_answers", 0)),
        ("Réflexes sauvegardés" if fr else "Reflexes saved", len(portrait.get("reflexes", []))),
    ]
    _rows(ws, 5, info, [26, 46])

    r = 5 + len(info) + 2
    ws.cell(row=r, column=1, value=("Vos droits" if fr else "Your rights")).font = Font(name=FONT, bold=True, size=11, color=INK)
    ws.cell(row=r + 1, column=1, value=(
        "Ce fichier contient toutes les données que Certifizer conserve sur vous. "
        "Vous pouvez y accéder, les rectifier, les emporter ailleurs, ou demander leur effacement. "
        "Conformément au RGPD et à la loi algérienne 18-07. Contact : contact@certifizer.app"
        if fr else
        "This file contains every piece of data Certifizer holds about you. You may access, correct, "
        "port or erase it. Under GDPR and Algerian law 18-07. Contact: contact@certifizer.app"))
    ws.cell(row=r + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 3, end_column=2)

    # ---------- 2. Ma progression ----------
    ws = wb.create_sheet("Ma progression" if fr else "My progress")
    ws.sheet_view.showGridLines = False
    _header(ws, 1, ["Domaine" if fr else "Area", "Score", ("Réponses" if fr else "Answers"),
                    ("État" if fr else "State"), ("S'appuie sur" if fr else "Depends on")], INK)
    STATE = {"acquired": "acquis" if fr else "acquired",
             "in_progress": "en cours" if fr else "in progress",
             "fragile": "fragile",
             "untouched": "vous attend" if fr else "awaits you"}
    data = [[n["label"], f"{round(n['score'] * 100)} %", n["attempts"],
             STATE.get(n["state"], n["state"]), ", ".join(n.get("depends_on", [])) or "—"]
            for n in portrait.get("nodes", [])]
    _rows(ws, 2, data, [22, 10, 12, 16, 30])

    # ---------- 3. Mon chemin critique ----------
    ws = wb.create_sheet("Chemin critique" if fr else "Critical path")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Le chemin critique — la chaîne qui commande votre date" if fr else "The critical path — the chain that governs your date"
    ws["A1"].font = Font(name=FONT, bold=True, size=12, color=INK)
    ws["A3"] = portrait.get("reading", "").replace("**", "")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A3:D6")
    ws.column_dimensions["A"].width = 28
    for c in "BCD":
        ws.column_dimensions[c].width = 22
    _header(ws, 8, ["#", ("Domaine (dans l'ordre)" if fr else "Area (in order)"), "Score", ("État" if fr else "State")], AMBER)
    by_id = {n["area"]: n for n in portrait.get("nodes", [])}
    cp = [[i + 1, by_id.get(a, {}).get("label", a),
           f"{round(by_id.get(a, {}).get('score', 0) * 100)} %",
           STATE.get(by_id.get(a, {}).get("state", ""), "")]
          for i, a in enumerate(portrait.get("critical_path", []))]
    _rows(ws, 9, cp, [6, 28, 12, 16])

    # ---------- 4. Mes réflexes ----------
    ws = wb.create_sheet("Mes réflexes" if fr else "My reflexes")
    ws.sheet_view.showGridLines = False
    _header(ws, 1, [("Réflexe (vos mots)" if fr else "Reflex (your words)"),
                    ("Depuis quel siège" if fr else "From which seat"), "Date"], TEAL)
    refl = [[x["text"], x.get("seat_label", ""), (x.get("at", "") or "")[:10]]
            for x in portrait.get("reflexes", [])]
    _rows(ws, 2, refl or [[("Aucun réflexe sauvegardé pour l'instant." if fr else "No reflexes saved yet."), "", ""]],
          [70, 20, 14])

    # ---------- 5. Ma trajectoire ----------
    ws = wb.create_sheet("Ma trajectoire" if fr else "My trajectory")
    ws.sheet_view.showGridLines = False
    _header(ws, 1, ["Date", ("Préparation" if fr else "Readiness")], INK)
    traj = [[t["day"], f"{round(t['readiness'] * 100)} %"] for t in portrait.get("trajectory", [])]
    _rows(ws, 2, traj or [[("Pas encore d'historique." if fr else "No history yet."), ""]], [16, 16])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
